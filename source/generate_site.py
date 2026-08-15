#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
海南话字典 — static site generator.
Reads the master Excel, writes a complete browsable website.

Usage:  python3 generate_site.py [--master PATH] [--ci PATH] [--out DIR]
"""
import argparse, json, os, re, shutil, sys, unicodedata
from collections import defaultdict, OrderedDict
import openpyxl

SUP        = "¹²³⁴⁵⁶⁷⁸"
SUP2NUM    = {c: str(i + 1) for i, c in enumerate(SUP)}
CIRCLED    = "㊀㊁㊂㊃㊄"
FILLED     = "❶❷❸❹❺❻❼❽❾❿⓫⓬"
SENSE_NUM  = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫"
MARKERS    = OrderedDict([("读", "文读"), ("话", "白读"), ("俗", "俗读"),
                          ("误", "误读"), ("又", "又读"), ("旧", "旧读")])
AUDIO_EXT  = ".mp3"        # None = use the filename exactly as given in the sheet

# ----------------------------------------------------------------------------- helpers
def strip_tone_marks(s):
    """gēng -> geng ; also drops ü diacritic to v-free form."""
    if not s:
        return ""
    out = unicodedata.normalize("NFD", s)
    out = "".join(c for c in out if unicodedata.category(c) != "Mn")
    return out.replace("ü", "u").lower()

def rom_plain(rom):
    """geng¹ -> geng ; ddo⁵ -> ddo"""
    return "".join(c for c in rom if c not in SUP)

def rom_digit(rom):
    """geng¹ -> geng1"""
    return "".join(SUP2NUM.get(c, c) for c in rom)

def first_syllable(rom):
    m = re.match(r"([a-z]+[" + SUP + r"]?)", rom)
    return m.group(1) if m else rom

def tone_of(rom):
    for c in rom:
        if c in SUP:
            return int(SUP2NUM[c])
    return 0

def esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))

# ----------------------------------------------------------------------------- column F
class Block:
    __slots__ = ("rom", "markers", "scopes", "marker", "homo", "paren", "alt_rom", "raw")
    def __init__(self):
        self.rom, self.markers, self.scopes = "", set(), set()
        self.marker, self.homo, self.paren, self.alt_rom, self.raw = None, "", "", "", ""

def _expand_scopes(text):
    """❶一❼ -> {1..7} ; ❶❸ -> {1,3}"""
    out, i = set(), 0
    nums = [(m.start(), FILLED.index(m.group()) + 1)
            for m in re.finditer("[" + FILLED + "]", text)]
    used = set()
    for idx, (pos, n) in enumerate(nums):
        if idx in used:
            continue
        nxt = nums[idx + 1] if idx + 1 < len(nums) else None
        if nxt and text[pos + 1:nxt[0]].strip() in ("一", "—", "-", "~", "～"):
            out.update(range(n, nxt[1] + 1))
            used.add(idx + 1)
        else:
            out.add(n)
    return out

def parse_f(cell):
    """Column F -> [Block]"""
    blocks = []
    if not cell:
        return blocks
    raw_chunks = re.findall(r"[\[〔]([^\]〕]*)[\]〕]", str(cell))
    chunks = []
    for rc in raw_chunks:                      # one bracket may hold two readings: "gan⁴干⁴，❷gan⁶干⁶"
        parts = [x for x in re.split(r"[，,；;/／]", rc) if x.strip()]
        carry = "".join(c for c in parts[0] if c in CIRCLED) if parts else ""
        for i, part in enumerate(parts):
            if i and carry and not any(c in CIRCLED for c in part):
                part = carry + part
            chunks.append(part)
    for chunk in chunks:
        b = Block(); b.raw = chunk
        b.markers = {c for c in chunk if c in CIRCLED}
        b.scopes = _expand_scopes(chunk)
        body = chunk
        # parentheses -> spelling gloss or an alternate reading
        parens = re.findall(r"[（(]([^）)]*)[）)]", body)
        body = re.sub(r"[（(][^）)]*[）)]", "", body)
        if parens:
            b.paren = parens[0]
            alt = re.findall(r"[a-z]+[" + SUP + r"]?", b.paren)
            if alt:
                b.alt_rom = "".join(alt)
        # marker word: sits before the latin
        head = re.split(r"[a-z]", body, 1)[0]
        for k, v in MARKERS.items():
            if k in head:
                b.marker = v
                break
        toks = re.findall(r"[a-z]+[" + SUP + r"]?", body)
        b.rom = "".join(toks)
        rest = re.sub(r"[a-z]+[" + SUP + r"]?", "", body)
        rest = re.sub(r"[" + CIRCLED + FILLED + r"\s，,]", "", rest)
        b.homo = "".join(c for c in rest if c not in MARKERS)
        if b.rom:
            blocks.append(b)
    return blocks

# ----------------------------------------------------------------------------- definitions
class Unit:
    __slots__ = ("marks", "pinyin", "senses")
    def __init__(self, marks=None, pinyin=None):
        self.marks, self.pinyin, self.senses = set(marks or ()), pinyin, []

def parse_def(text):
    """Definition cell -> [Unit]. A line like "㈚㈛ jiā / jiá" sets the marker scope;
    any 〔pinyin〕 blocks that follow inherit it."""
    units = []
    if not text:
        return units
    cur, cur_marks, fresh = None, set(), False
    for raw in str(text).split("\n"):
        line = raw.strip()
        if not line:
            continue
        if line[0] in CIRCLED:
            cur_marks = {c for c in line if c in CIRCLED}
            head = "".join(c for c in line if c not in CIRCLED).strip()
            cur = Unit(cur_marks, head or None)
            units.append(cur); fresh = True; continue
        m = re.match(r"^[〔\[]([^〕\]]+)[〕\]]\s*$", line)
        if m:
            py = m.group(1).strip()
            if cur is not None and not cur.senses and fresh:
                cur.pinyin = py
            else:
                cur = Unit(cur_marks, py); units.append(cur)
            fresh = False; continue
        if cur is None:
            cur = Unit(); units.append(cur)
        if line[0] in SENSE_NUM:
            cur.senses.append(line[1:].strip())
        elif cur.senses:
            cur.senses[-1] += " " + line
        else:
            cur.senses.append(line)
    return [u for u in units if u.senses]

def select_units(units, marks):
    """Units belonging to a set of ㈚㈛ markers (all units when nothing is marked)."""
    if not units:
        return []
    if marks:
        sel = [u for u in units if u.marks & marks]
        if sel:
            return sel
        if any(u.marks for u in units):
            return []
    return units

# ----------------------------------------------------------------------------- entry model
class Entry:
    def __init__(self, ch, rom, page, pinyin, variants):
        self.ch, self.rom, self.page = ch, rom, page
        self.pinyin, self.variants = pinyin, variants
        self.homo = ""
        self.paren = ""      # 反切 spelling gloss, or a parenthetical alternate reading
        self.paren_rom = ""
        self.units = []          # [(pinyin_or_None, [(cn, en, [chips])])]
        self.head_alts = []      # [(marker, rom)]
        self.pointer = None      # rom of the canonical entry
        self.others = []         # 另见 roms
        self.ci = []             # 常用词: [{w,rom,cn,en}]
    @property
    def tone(self):
        return tone_of(self.rom)
    @property
    def syllable(self):
        return rom_plain(first_syllable(self.rom))
    @property
    def anchor(self):
        return "e-%s-%s" % (rom_digit(self.rom), self.ch)

def build_entries(row, ci_map):
    (_, page, ch, pinyin, fcell, cn, _forms, _pb, _st, _cat, _act, en, *_rest) = row
    blocks   = parse_f(fcell)
    if not blocks:
        return [], "no reading parsed"
    cn_units = parse_def(cn)
    en_units = parse_def(en)
    primary  = [b for b in blocks if not b.marker]
    marked   = [b for b in blocks if b.marker]
    if not primary:                       # everything is marked: promote the first
        primary, marked = blocks[:1], blocks[1:]

    # group primary blocks by identical romanization (a merge)
    groups = OrderedDict()
    for b in primary:
        groups.setdefault(b.rom, []).append(b)

    flag = None

    entries = []
    for rom, bl in groups.items():
        e = Entry(ch, rom, page, pinyin, None)
        e.homo = next((b.homo for b in bl if b.homo), "")
        pb = next((b for b in bl if b.paren), None)
        if pb:
            e.paren, e.paren_rom = pb.paren, pb.alt_rom
        marks = set().union(*[b.markers for b in bl]) if bl else set()
        sel_cn = select_units(cn_units, marks)
        sel_en = select_units(en_units, marks)
        if not sel_cn and cn_units:
            sel_cn, sel_en = cn_units, en_units
            flag = "markers in column F not found in the definition"
        e.pinyin = " / ".join([u.pinyin for u in sel_cn if u.pinyin]) or \
                   re.sub("[" + CIRCLED + "]", "", str(pinyin or "")).replace("\n", " / ").strip()
        for i, u in enumerate(sel_cn):
            eu = sel_en[i] if i < len(sel_en) else None
            senses = []
            for j, s in enumerate(u.senses):
                senses.append([s, (eu.senses[j] if eu and j < len(eu.senses) else ""), []])
            e.units.append([u.pinyin, senses])
        entries.append(e)

    # several unmarked readings covering the same senses -> pointer entries
    if len(entries) > 1 and all(not b.markers for b in primary):
        canon = entries[0]
        for e in entries[1:]:
            e.units, e.pointer = [], canon.rom

    # attach marked readings
    for b in marked:
        targets = [e for e in entries
                   if not b.markers or any(m in b.markers for m in
                       set().union(*[x.markers for x in groups[e.rom]]) or {""})] or entries
        for e in targets:
            if b.scopes and e.units:
                flat = [s for _, ss in e.units for s in ss]
                for n in sorted(b.scopes):
                    if 1 <= n <= len(flat):
                        flat[n - 1][2].append((b.marker, b.rom))
                    else:
                        flag = "sense scope out of range"
            else:
                if (b.marker, b.rom) not in e.head_alts:
                    e.head_alts.append((b.marker, b.rom))

    roms = [e.rom for e in entries]
    for e in entries:
        e.others = [r for r in roms if r != e.rom]
    return entries, flag

# ----------------------------------------------------------------------------- rendering
GLOSS_CN = {
 "书": "书面语。多见于书面或文言，口语少用。",
 "方": "方言字，或此义项是方言地区的用法。",
 "旧": "旧时用法。今已少用或不用。",
 "古": "古代用字，或此义项是古代的用法。",
 "外": "近代外来语。沿用已久的外来语一般不注来源。",
 "俗": "俗语。通俗的口头说法，不用于正式场合。",
 "话": "海南话特有的说法，普通话无此用法。",
 "引": "由原义引申出来的意义。",
 "喻": "由比喻形成的意义。",
 "罕": "罕用。极少使用。",
 "lit.": "Literary. Chiefly written or classical style; rare in speech.",
 "dial.": "Dialectal. Used only in certain dialect areas.",
 "former": "Former usage. Little or no longer used today.",
 "arch.": "Archaic. Found in old texts; not current in modern Chinese.",
 "loan": "Loanword. Borrowed or transliterated from a foreign language.",
 "colloq.": "Colloquial. Everyday spoken usage; not for formal contexts.",
 "Hain.": "Specific to spoken Hainanese; no equivalent usage in Mandarin.",
 "ext.": "Extended sense. Developed out of the word's basic meaning.",
 "fig.": "Figurative sense. Used metaphorically rather than literally.",
 "rare": "Rare. Very seldom used.",
}
MARK_TIP = {
 "文读": "文读（读音）。读书、书面语、成语、专名里用的读法，跟日常口语的说法不同。",
 "白读": "白读（话音）。日常说话时用的读法。",
 "俗读": "俗读。通行的通俗读法，虽不合正音，但一般人都这么念。",
 "误读": "误读。本不合正音，却已约定俗成、广泛通行的读法；也包括训读——借用同义字的音来念本字。",
 "又读": "又读。同一义项下并存的另一种读法，两读皆可。",
 "旧读": "旧读。旧时的读法，今已改读。",
}
NUM_CH = "　①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"

def mark_labels(text):
    def rep(m):
        k = m.group(1)
        tip = GLOSS_CN.get(k, "")
        return ('<span class="gloss" data-tip="%s">〈%s〉</span>' % (esc(tip), esc(k))
                if tip else m.group(0))
    return re.sub(r"〈([^〉]+)〉", rep, esc(text))

def reading_html(rom, size="", audio=None):
    """Renders each syllable with its own tone badge: ddeng¹ddang¹ -> ddeng(1)ddang(1)."""
    parts = re.findall(r"([a-z]+)([" + SUP + r"]?)", rom)
    out = "".join(esc(syl) + ('<i>%s</i>' % SUP2NUM[t] if t else "")
                  for syl, t in parts if syl)
    return '<span class="hai %s">%s</span>' % (size, out or esc(rom_plain(rom)))

def audio_btn(rom, audio_map, small=False):
    fn = audio_map.get(rom)
    if not fn:
        # compound reading: concatenate the single-syllable filenames (ddeng1ddang1.mp3)
        syls = ["".join(t) for t in re.findall(r"([a-z]+)([" + SUP + r"]?)", rom) if t[0]]
        if len(syls) > 1 and all(s_ in audio_map for s_ in syls):
            ext = os.path.splitext(audio_map[syls[0]])[1]
            fn = "".join(os.path.splitext(audio_map[s_])[0] for s_ in syls) + ext
    if not fn:
        return ""
    return ('<button class="play%s" data-src="audio/%s">%s音频</button>'
            % (" sm" if small else "", esc(fn), SPK))

SPK = ('<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" '
       'stroke-linecap="round" stroke-linejoin="round"><polygon points="11 5 6 9 2 9 2 15 '
       '6 15 11 19 11 5"></polygon><path d="M15.5 8.5a5 5 0 0 1 0 7"></path>'
       '<path d="M19 5a9 9 0 0 1 0 14"></path></svg>')

def chip(rom, ch, offpage=False):
    return ('<a class="rchip" href="%s.html#%s">%s<i>%d</i>%s</a>'
            % (rom_plain(first_syllable(rom)), "e-%s-%s" % (rom_digit(rom), ch),
               esc(rom_plain(rom)), tone_of(rom),
               '<span class="off">↗</span>' if offpage else ""))

def entry_html(e, audio_map, new_tone):
    h = ['<div class="entry%s" id="%s" data-tone="%d">' % (" newtone" if new_tone else "",
                                                           e.anchor, e.tone)]
    h.append('<div class="head"><div class="box"><span>%s</span></div><div class="labels">')
    h[-1] = h[-1] % esc(e.ch)
    h.append('<div class="r-hai"><span class="tag-hai tip" data-tip="海南话读音，以文昌县城音为代表；'
             '个别字另注其他地方的读法。">海南话</span>%s%s' % (reading_html(e.rom),
                                                        audio_btn(e.rom, audio_map)))
    if e.homo:
        h.append('<span class="homo tip" data-tip="这个字的海南话读音跟这个字完全相同。">音同「%s」</span>'
                 % esc(e.homo))
    if e.paren and not e.paren_rom:
        h.append('<span class="fanqie tip" data-tip="反切：取上一个字的声母、下一个字的韵母，'
                 '拼成这个读音。">（%s）</span>' % esc(e.paren))
    if e.paren_rom:
        h.append('<span class="altread"><span class="lbl tip" data-tip="同一读音的另一种写法，'
                 '两者通用。">亦作</span>%s%s</span>'
                 % (reading_html(e.paren_rom, "sm"), audio_btn(e.paren_rom, audio_map, True)))
    for mk, rom in e.head_alts:
        h.append('<span class="altread"><span class="lbl tip" data-tip="%s">%s</span>%s%s</span>'
                 % (esc(MARK_TIP.get(mk, "")), esc(mk), reading_html(rom, "sm"),
                    audio_btn(rom, audio_map, True)))
    h.append("</div>")
    if e.pinyin:
        h.append('<div class="r-man"><span class="tag-man tip" data-tip="普通话读音，用汉语拼音标注。">'
                 '普通话</span><span class="man">%s</span></div>'
                 % esc(str(e.pinyin).replace("\n", " / ")))
    if e.others:
        chips = "".join(chip(r, e.ch, rom_plain(first_syllable(r)) != e.syllable)
                        for r in e.others)
        h.append('<div class="r-see"><span class="tag-see tip" data-tip="这个字的其他海南话读音。'
                 '点一下就跳到那个读音的条目。">另见</span>%s</div>' % chips)
    h.append("</div></div>")

    h.append('<div class="sec-h">字义解释</div>')
    if e.pointer:
        h.append('<div class="pointer">此读音不另立释义 — 释义见 %s</div>'
                 % chip(e.pointer, e.ch,
                        rom_plain(first_syllable(e.pointer)) != e.syllable))
    elif not e.units:
        h.append('<div class="pointer">释义待补</div>')
    else:
        n = 1
        multi = len([1 for p, _ in e.units if p]) > 1
        for py, senses in e.units:
            if py and multi:
                h.append('<div class="subblock">〔%s〕</div>' % esc(py))
            for cn, en, chips_ in senses:
                bits = "".join(
                    '<span class="altread"><span class="lbl tip" data-tip="%s">%s</span>%s%s</span>'
                    % (esc(MARK_TIP.get(mk, "")), esc(mk), reading_html(rm, "sm"),
                       audio_btn(rm, audio_map, True))
                    for mk, rm in chips_)
                h.append('<div class="sense">%s %s%s%s</div>'
                         % (NUM_CH[n] if n < len(NUM_CH) else "(%d)" % n,
                            mark_labels(cn), bits,
                            '<span class="en">%s</span>' % mark_labels(en) if en else ""))
                n += 1
    h.append('<div class="sec-h">常用词</div>')
    if e.ci:
        for w in e.ci:
            h.append('<div class="ci"><div class="w">%s</div>%s%s<div class="cidef">%s%s</div></div>'
                     % (esc(w["w"]),
                        reading_html(w["rom"], "sm") if w["rom"] else "",
                        audio_btn(w["rom"], audio_map, True) if w["rom"] else "",
                        mark_labels(w["cn"]),
                        '<span class="en">%s</span>' % mark_labels(w["en"]) if w["en"] else ""))
    else:
        h.append('<div class="ci-empty">待补充 — 此处留给日后收录的词条</div>')
    h.append("</div>")
    return "".join(h)

# ----------------------------------------------------------------------------- page shell
def page(title, body, depth_note="", extra="", syl=None):
    attrs = ' class="has-pn" data-syl="%s"' % esc(syl) if syl else ""
    return """<!DOCTYPE html>
<html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>%s — 海南话字典</title>
<link rel="stylesheet" href="site.css"></head><body%s>
%s
<footer>海南话字典 · 由主表自动生成，请勿直接改动网页。%s</footer>
<script src="site.js"></script>%s</body></html>""" % (esc(title), attrs, body, depth_note, extra)

def panel_shell(n_chars):
    """Empty panel. site.js fetches suoyin_panel.html once and fills it — inlining the
    index into all 517 pages would weigh several hundred megabytes."""
    return ('<div class="pn-scrim"></div>'
            '<aside class="pn" id="pn">'
            '<div class="pn-top"><span class="t">音序索引</span>'
            '<span class="n">%d 字</span>'
            '<a class="pn-home" href="index.html">首页</a>'
            '<button class="pn-x" type="button" aria-label="收起索引">&times;</button></div>'
            '<div class="pn-filter"><input id="pn-q" type="search" '
            'placeholder="筛选音节或汉字" autocomplete="off"></div>'
            '<div class="pn-body" id="pn-body">'
            '<div class="pn-loading">索引载入中…</div></div></aside>' % n_chars)

def panel_fragment(by_syl, letters):
    """The one-column 音序索引, written once and shared by every page."""
    h = []
    for k, syls in letters.items():
        h.append('<div class="pn-L"><div class="pn-letter">%s</div>' % esc(k))
        for s in syls:
            h.append('<div class="pn-g" data-syl="%s">'
                     '<div class="pn-gh">%s<em>(%s)</em><span class="ct">%d</span></div>'
                     % (esc(s), esc(s.upper()), esc(s), len(by_syl[s])))
            for e in by_syl[s]:
                h.append('<a class="pn-r" data-c="%s" href="%s.html#%s">'
                         '<span class="z">%s</span><i>%d</i><span class="dots"></span>'
                         '<span class="pg">%s</span></a>'
                         % (esc(e.ch), e.syllable, e.anchor, esc(e.ch), e.tone,
                            esc(str(e.page or ""))))
            h.append('</div>')
        h.append('</div>')
    return "".join(h)

def syllable_page(syl, entries, audio_map, all_syls, n_chars):
    tones = sorted({e.tone for e in entries})
    chips = "".join('<a href="#" data-t="%d">%s<sup>%d</sup></a>' % (t, esc(syl), t)
                    for t in tones)
    body = [panel_shell(n_chars)]
    body.append('<div class="bar"><div class="bar-in">'
                '<div class="bar-row">'
                '<button class="pn-toggle" type="button">≡ 索引</button>'
                '<div class="syl" id="bar-syl">%s<small>音节</small></div>'
                '<input id="q" class="qbox" placeholder="搜字 · 海南话 · 拼音" autocomplete="off">'
                '</div><div class="jump" id="jump">%s</div><div id="res" class="res"></div>'
                '</div></div>' % (esc(syl), chips))
    body.append('<div class="wrap">')
    body.append('<div class="ss-gate" id="ss-gate-top"></div>')
    body.append('<div id="stream"><section class="ss" data-syl="%s">' % esc(syl))
    body.append('<div class="ss-head"><span class="ss-syl">%s<small>音节</small></span>'
                '<span class="ss-n">%d 读音</span></div>' % (esc(syl), len(entries)))
    prev = None
    for e in entries:
        body.append(entry_html(e, audio_map, e.tone != prev))
        prev = e.tone
    body.append('</section></div>')
    body.append('<div class="ss-gate" id="ss-gate"></div>')
    body.append('<div class="ss-load" id="ss-note"></div>')
    # prev/next remain as the no-JS fallback; site.js removes them once streaming is live
    i = all_syls.index(syl)
    prev_s = all_syls[i - 1] if i else None
    next_s = all_syls[i + 1] if i + 1 < len(all_syls) else None
    nav = []
    if prev_s: nav.append('<a href="%s.html">← %s</a>' % (prev_s, prev_s))
    nav.append('<a href="index.html">全部音节</a>')
    if next_s: nav.append('<a href="%s.html">%s →</a>' % (next_s, next_s))
    body.append('<div class="pagenav">%s</div></div>' % " · ".join(nav))
    return page(syl, "".join(body), syl=syl)

# ----------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--master", default="Hainanese_dictionary_master_with_鼾.xlsx")
    ap.add_argument("--ci", default="词.xlsx")
    ap.add_argument("--out", default="site")
    ap.add_argument("--assets", default="assets")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.master, read_only=True, data_only=True)
    rows = list(wb["All Rows"].iter_rows(min_row=2, values_only=True))
    audio_map = {}
    for r in wb["Unique Romanizations"].iter_rows(min_row=2, values_only=True):
        if r[0] and r[5]:
            fn = str(r[5])
            if AUDIO_EXT:
                fn = os.path.splitext(fn)[0] + AUDIO_EXT
            audio_map[str(r[0])] = fn

    all_names = {str(r[2]) for r in rows if r[2]}
    singles = {c for c in all_names if len(c) == 1}
    word_rows = [r for r in rows if r[2] and len(str(r[2])) > 1]

    # every multi-character headword becomes a 常用词 of one of its characters
    ci_map, orphan_rows, host_stat = defaultdict(list), [], {"first": 0, "second": 0}
    for r in word_rows:
        w = str(r[2])
        pos = next((i for i, c in enumerate(w) if c in singles), None)
        if pos is None:
            orphan_rows.append(r)               # neither character is a headword
            continue
        host_stat["first" if pos == 0 else "second"] += 1
        blks = parse_f(r[4])
        syls = re.findall(r"[a-z]+[" + SUP + r"]?", blks[0].rom) if blks else []
        host_rom = syls[pos] if pos < len(syls) else (syls[0] if syls else "")
        rom = "".join(syls)                     # whole-word reading, e.g. ddeng¹ddang¹
        ci_map[w[pos]].append({"w": w, "rom": rom, "host_rom": host_rom,
                               "cn": str(r[5] or "").replace("\n", " ").strip(),
                               "en": str(r[11] or "").replace("\n", " ").strip()})
    print("常用词: %d placed (%d under the first character, %d under the second), "
          "%d kept as their own entries"
          % (sum(len(v) for v in ci_map.values()), host_stat["first"], host_stat["second"],
             len(orphan_rows)))

    entries, flags = [], []
    for r in rows:
        if r[2] and len(str(r[2])) > 1 and r not in orphan_rows:
            continue                            # 联绵词 live under 常用词, not as headwords
        try:
            es, flag = build_entries(r, ci_map)
        except Exception as exc:                       # never let one row kill the build
            es, flag = [], "parse error: %s" % exc
        if flag:
            flags.append((r[1], r[2], flag))
        entries.extend(es)

    by_char = defaultdict(list)
    for e in entries:
        by_char[e.ch].append(e)
    for host, words in ci_map.items():
        targets = by_char.get(host, [])
        if not targets:
            continue
        for w in words:
            match = [e for e in targets if e.rom == w["host_rom"]] or targets
            for e in match:
                e.ci.append(w)

    by_syl = defaultdict(list)
    for e in entries:
        by_syl[e.syllable].append(e)
    for k in by_syl:
        by_syl[k].sort(key=lambda e: (e.tone, int(e.page or 0), e.ch))
    all_syls = sorted(by_syl)

    out = a.out
    if os.path.isdir(out):
        shutil.rmtree(out)
    os.makedirs(out)
    os.makedirs(os.path.join(out, "audio"), exist_ok=True)
    for f in ("site.css", "site.js"):
        shutil.copy(os.path.join(a.assets, f), os.path.join(out, f))

    n_chars = len({e.ch for e in entries})
    letters_ix = OrderedDict()
    for s in all_syls:
        letters_ix.setdefault(s[0].upper(), []).append(s)
    with open(os.path.join(out, "suoyin_panel.html"), "w", encoding="utf-8") as fh:
        fh.write(panel_fragment(by_syl, letters_ix))

    for syl in all_syls:
        with open(os.path.join(out, syl + ".html"), "w", encoding="utf-8") as fh:
            fh.write(syllable_page(syl, by_syl[syl], audio_map, all_syls, n_chars))

    # search index
    idx, seen = [], set()
    for e in entries:
        url = "%s.html#%s" % (e.syllable, e.anchor)
        pin = strip_tone_marks(str(e.pinyin or "").replace("\n", " "))
        idx.append({"c": e.ch, "r": rom_plain(e.rom), "t": e.tone,
                    "d": rom_digit(e.rom), "p": pin, "u": url})
        alts = list(e.head_alts) + [(m, r) for _, ss in e.units for s_ in ss for m, r in s_[2]]
        for mk, rom in alts:
            if (e.ch, rom) in seen:
                continue
            seen.add((e.ch, rom))
            idx.append({"c": e.ch, "r": rom_plain(rom), "t": tone_of(rom),
                        "d": rom_digit(rom), "p": pin, "u": url, "m": mk})
    with open(os.path.join(out, "index.json"), "w", encoding="utf-8") as fh:
        json.dump(idx, fh, ensure_ascii=False, separators=(",", ":"))

    # home page
    letters = defaultdict(list)
    for s in all_syls:
        letters[s[0]].append(s)
    blocks = "".join(
        '<section class="ix"><h3>%s</h3><div class="sylgrid">%s</div></section>'
        % (k.upper(), "".join('<a href="%s.html">%s<em>%d</em></a>'
                              % (s, s, len(by_syl[s])) for s in letters[k]))
        for k in sorted(letters))
    home = ('<div class="bar"><div class="bar-in"><div class="bar-row">'
            '<div class="syl">海南话字典<small>hainanese dictionary</small></div>'
            '<input id="q" class="qbox" placeholder="搜字 · 海南话 · 拼音" autocomplete="off">'
            '</div><div id="res" class="res"></div></div></div>'
            '<div class="wrap"><p class="lede">共 %d 个字头、%d 条读音、%d 个音节。'
            '按海南话读音排列；同一音节内先按声调，再按原书页序。</p>%s</div>'
            % (len({e.ch for e in entries}), len(entries), len(all_syls), blocks))
    with open(os.path.join(out, "index.html"), "w", encoding="utf-8") as fh:
        fh.write(page("首页", home))

    # build report
    with open(os.path.join(out, "_build_report.txt"), "w", encoding="utf-8") as fh:
        fh.write("entries: %d\ncharacters: %d\nsyllables: %d\nflagged rows: %d\n\n"
                 % (len(entries), len({e.ch for e in entries}), len(all_syls), len(flags)))
        for p, c, f in flags:
            fh.write("p%s\t%s\t%s\n" % (p, c, f))

    print("pages: %d  entries: %d  chars: %d  flagged: %d"
          % (len(all_syls) + 1, len(entries), len({e.ch for e in entries}), len(flags)))

if __name__ == "__main__":
    main()
